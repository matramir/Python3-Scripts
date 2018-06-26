#! python3
# aprendolibre2.py third iteration of a script made for a friend that extracts the liks from several raw html .txt files
# Usage: you should have on the same pwd .txt files with raw html code with the following format: AprendoRAW<number from 1 to n>
import os, bs4, re
import openpyxl

def excelCreation(rawN):
    url = 'http://www.aprendolibre.cl' #this can be changed to fit another purpose in the future
    selector = 'a' #this can be changed to fit another purpose
    #regex definition
    groupN = re.compile(r'''
    (http://www.aprendolibre.cl/materiales/)      #subfolder
    ([0-9]+)          #groupN

    ''',re.VERBOSE)
    #excel  document creation
    wb = openpyxl.Workbook()
    skipStr = ['None','Abrir','Editar cuenta','Cerrar sesión']#this can be modified to fit different contents that the user do not want displayed on the list


    #endfile .txt creation
    fileName = 'aprendo'+str(rawN)+'.txt'
    endFile = os.path.join(os.path.abspath('.'),fileName)
    faprendo = open(endFile, 'w')#debugging purposes

    for repetition in range(rawN):

        #import data from the raw html file
        htmlRaw = open(os.path.join(os.path.abspath('.'),'AprendoRAW'+str(repetition+1)+'.txt'))
        htmlText = htmlRaw.read()
        htmlRaw.close()
        #print(htmlText) #debugging purposes

        #beautiful soup object creation
        soup = bs4.BeautifulSoup(htmlText, 'html.parser')
        #print(soup.prettify()) #debugging purposes
        elems = soup.find_all(selector)

        #excel sheet definition
        sheet = wb.active
        wb.create_sheet('Scrap'+str(repetition+1),0)
        ws = wb['Scrap'+str(repetition+1)]
        cellN = 2

        for link in elems:
            linkUrl = url+str(link.get('href'))
            linkStr = str(link.string)
            faprendo.write('\n'+linkUrl+' - '+ linkStr)
            print(linkUrl+' - '+ linkStr)

            if linkStr not in skipStr:
                #cell definition
                cell = 'A'+str(cellN)
                gCell = 'B'+str(cellN)
                #Writing of de link column
                ws[cell] = '=HYPERLINK("{}", "{}")'.format(linkUrl, linkStr)
                print(ws[cell].value)

                gs = groupN.search(linkUrl)
                gResult = gs.group(2)
                ws[gCell] = gResult
                print(ws[gCell].value)

                cellN += 1

            else:
                continue

    faprendo.close()
    wb.save('alresults.xlsx')

print('Please enter a number of aprendoRaw files:')
try:
    excelCreation(int(input()))
except:
    print('Something went wrong, please try again.\n(Quicktip: The raw html files should be named in the following format:\n\"AprendoRAW<number from 1 to n>)')
